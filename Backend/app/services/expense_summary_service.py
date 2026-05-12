# -*- coding: utf-8 -*-
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from typing import Literal
import csv
import io
from sqlalchemy.orm import Session

from app.enums import CurrencyEnum
from app.repositories import ExpenseRepository
from app.utils.export_localization import (
    CSV_EXPORT_COLUMNS,
    EXPORT_ALL_SECTIONS,
    EXPORT_ALLOWED_SECTIONS,
    EXPORT_DEFAULT_SECTIONS,
    EXPORT_LOCALIZATION,
    ExportFormat,
    ExportSection,
)
from .group_service import GroupService

SummaryScope = Literal["all", "personal", "group"]
TrendGranularity = Literal["daily", "weekly", "monthly"]
DrilldownSortBy = Literal["expense_date", "amount", "created_at"]
SortOrder = Literal["asc", "desc"]


class ExpenseSummaryService:
    def __init__(self, db: Session):
        self.expense_repo = ExpenseRepository(db)
        self.group_service = GroupService(db)

    @staticmethod
    def _to_decimal(value) -> Decimal:
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value or 0))

    def _resolve_date_range(self, range_value: str) -> tuple[date, date]:
        today = date.today()
        if range_value == "current_week":
            start = today - timedelta(days=today.weekday())
            end = today
        elif range_value == "previous_week":
            end = today - timedelta(days=today.weekday() + 1)
            start = end - timedelta(days=6)
        elif range_value == "current_month":
            start = date(today.year, today.month, 1)
            end = today
        elif range_value == "previous_month":
            if today.month == 1:
                start = date(today.year - 1, 12, 1)
                end = date(today.year - 1, 12, 31)
            else:
                start = date(today.year, today.month - 1, 1)
                import calendar
                _, last_day = calendar.monthrange(today.year, today.month - 1)
                end = date(today.year, today.month - 1, last_day)
        else:
            raise HTTPException(status_code=400, detail="range must be one of: current_week, previous_week, current_month, previous_month")
        return start, end

    def get_categories_by_range(
        self,
        user_id: int,
        range_value: str,
        currency: CurrencyEnum,
    ):
        date_from, date_to = self._resolve_date_range(range_value)
        date_from_dt = datetime.combine(date_from, time.min)
        date_to_dt = datetime.combine(date_to, time.max)

        personal_rows = self.expense_repo.get_personal_category_totals(
            user_id=user_id,
            date_from=date_from_dt,
            date_to=date_to_dt,
            currency=currency,
        )
        group_rows = self.expense_repo.get_group_category_totals(
            user_id=user_id,
            date_from=date_from_dt,
            date_to=date_to_dt,
            currency=currency,
        )

        category_map: dict[int, dict] = {}
        total_amount = Decimal("0")

        for row in personal_rows + group_rows:
            cat_id = int(row.category_id)
            current = category_map.setdefault(
                cat_id,
                {
                    "category_id": cat_id,
                    "category_name": row.category_name,
                    "amount": Decimal("0"),
                    "currency": currency,
                },
            )
            amount = Decimal(str(row.total_amount or 0))
            current["amount"] += amount
            total_amount += amount

        categories = sorted(
            category_map.values(),
            key=lambda item: item["amount"],
            reverse=True,
        )

        return {
            "currency": currency,
            "range": range_value,
            "total_amount": total_amount,
            "categories": categories,
        }

    def get_trend_by_range(
        self,
        user_id: int,
        range_value: str,
        currency: CurrencyEnum,
    ):
        date_from, date_to = self._resolve_date_range(range_value)
        date_from_dt = datetime.combine(date_from, time.min)
        date_to_dt = datetime.combine(date_to, time.max)

        personal_rows = self.expense_repo.get_personal_daily_totals(
            user_id=user_id,
            date_from=date_from_dt,
            date_to=date_to_dt,
            currency=currency,
        )
        group_rows = self.expense_repo.get_group_daily_totals(
            user_id=user_id,
            date_from=date_from_dt,
            date_to=date_to_dt,
            currency=currency,
        )

        day_map: dict[date, Decimal] = {}
        for row in personal_rows + group_rows:
            day = row.day if isinstance(row.day, date) else date.fromisoformat(str(row.day))
            day_map[day] = day_map.get(day, Decimal("0")) + Decimal(str(row.total_amount or 0))

        current = date_from
        points = []
        while current <= date_to:
            points.append({
                "date": current,
                "amount": day_map.get(current, Decimal("0")),
                "currency": currency,
            })
            current += timedelta(days=1)

        return {
            "currency": currency,
            "range": range_value,
            "date_from": date_from,
            "date_to": date_to,
            "points": points,
        }

    @staticmethod
    def _to_currency(value) -> CurrencyEnum:
        return value if isinstance(value, CurrencyEnum) else CurrencyEnum(str(value))

    @staticmethod
    def _normalize_export_locale(locale: str | None) -> str:
        if not locale:
            return "en"

        normalized = locale.strip().lower().replace("_", "-")
        if normalized.startswith("pl"):
            return "pl"

        return "en"

    @classmethod
    def _get_export_localization(cls, locale: str | None):
        locale_code = cls._normalize_export_locale(locale)
        return locale_code, EXPORT_LOCALIZATION[locale_code]

    @staticmethod
    def _format_export_date(value: datetime, locale: str) -> str:
        if locale == "pl":
            return value.strftime("%d.%m.%Y")

        return value.strftime("%Y-%m-%d")

    @staticmethod
    def _format_export_datetime(value: datetime, locale: str) -> str:
        if locale == "pl":
            return value.strftime("%d.%m.%Y %H:%M")

        return value.strftime("%Y-%m-%d %H:%M")

    @staticmethod
    def _format_export_amount(value: Decimal, locale: str) -> str:
        normalized = f"{Decimal(value):.2f}"
        if locale == "pl":
            return normalized.replace(".", ",")

        return normalized

    @staticmethod
    def _format_export_currency(value: CurrencyEnum | str) -> str:
        if isinstance(value, CurrencyEnum):
            return value.value

        raw = str(value)
        if raw.startswith("CurrencyEnum."):
            return raw.split(".", 1)[1]

        return raw

    @staticmethod
    def _normalize_day(value) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            return date.fromisoformat(value)
        raise ValueError("Unsupported day value")

    @staticmethod
    def _iter_days(date_from: date, date_to: date):
        current = date_from
        while current <= date_to:
            yield current
            current += timedelta(days=1)

    def _resolve_scope(self, scope: SummaryScope) -> SummaryScope:
        if scope not in ("all", "personal", "group"):
            raise HTTPException(status_code=400, detail="scope must be one of: all, personal, group")
        return scope

    @staticmethod
    def _resolve_trend_granularity(granularity: TrendGranularity) -> TrendGranularity:
        if granularity not in ("daily", "weekly", "monthly"):
            raise HTTPException(status_code=400, detail="granularity must be one of: daily, weekly, monthly")
        return granularity

    @staticmethod
    def _bucket_start_for_day(current_day: date, granularity: TrendGranularity) -> date:
        if granularity == "weekly":
            return current_day - timedelta(days=current_day.weekday())

        if granularity == "monthly":
            return current_day.replace(day=1)

        return current_day

    def _iter_trend_buckets(self, date_from: date, date_to: date, granularity: TrendGranularity):
        if granularity == "daily":
            yield from self._iter_days(date_from, date_to)
            return

        if granularity == "weekly":
            current = self._bucket_start_for_day(date_from, "weekly")
            last = self._bucket_start_for_day(date_to, "weekly")
            while current <= last:
                yield current
                current += timedelta(days=7)
            return

        # monthly
        current = self._bucket_start_for_day(date_from, "monthly")
        last = self._bucket_start_for_day(date_to, "monthly")
        while current <= last:
            yield current
            if current.month == 12:
                current = date(current.year + 1, 1, 1)
            else:
                current = date(current.year, current.month + 1, 1)

    def _aggregate_trend_buckets(
        self,
        trend_map: dict[CurrencyEnum, dict[date, dict[str, Decimal]]],
        period: tuple[date, date],
        granularity: TrendGranularity,
    ) -> dict[CurrencyEnum, dict[date, dict[str, Decimal]]]:
        if granularity == "daily":
            return trend_map

        aggregated: dict[CurrencyEnum, dict[date, dict[str, Decimal]]] = {}
        for currency_key, day_map in trend_map.items():
            currency_bucket = aggregated.setdefault(currency_key, {})
            for day_key, amounts in day_map.items():
                if day_key < period[0] or day_key > period[1]:
                    continue

                bucket_key = self._bucket_start_for_day(day_key, granularity)
                bucket_amounts = currency_bucket.setdefault(
                    bucket_key,
                    {"personal_amount": Decimal("0"), "group_amount": Decimal("0")},
                )
                bucket_amounts["personal_amount"] += amounts["personal_amount"]
                bucket_amounts["group_amount"] += amounts["group_amount"]

        return aggregated

    def _resolve_category_ids(
        self,
        category_id: int | None = None,
        category_ids: list[int] | None = None,
    ) -> list[int] | None:
        resolved_ids: list[int] = []

        if category_id is not None:
            resolved_ids.append(category_id)

        if category_ids:
            resolved_ids.extend(category_ids)

        if not resolved_ids:
            return None

        unique_ids = list(dict.fromkeys(resolved_ids))

        for current_category_id in unique_ids:
            if current_category_id <= 0:
                raise HTTPException(status_code=400, detail="category_ids must contain only positive integers")

        return unique_ids

    def _resolve_periods(
        self,
        date_from: date | None,
        date_to: date | None,
        compare_previous: bool,
    ) -> tuple[tuple[date, date], tuple[date, date] | None]:
        if (date_from is None) != (date_to is None):
            raise HTTPException(status_code=400, detail="date_from and date_to must be provided together")

        if date_from is None or date_to is None:
            today = date.today()
            date_from = date(today.year, today.month, 1)
            date_to = today

        if date_from > date_to:
            raise HTTPException(status_code=400, detail="date_from cannot be greater than date_to")

        current_period = (date_from, date_to)

        if not compare_previous:
            return current_period, None

        range_days = (date_to - date_from).days + 1
        previous_date_to = date_from - timedelta(days=1)
        previous_date_from = previous_date_to - timedelta(days=range_days - 1)
        previous_period = (previous_date_from, previous_date_to)

        return current_period, previous_period

    @staticmethod
    def _to_datetime_range(period: tuple[date, date]) -> tuple[datetime, datetime]:
        return (
            datetime.combine(period[0], time.min),
            datetime.combine(period[1], time.max),
        )

    def _validate_scope_group_filter(self, scope: SummaryScope, group_id: int | None, user_id: int) -> None:
        if scope == "personal" and group_id is not None:
            raise HTTPException(status_code=400, detail="group_id is not supported for personal scope")

        if group_id is not None:
            self.group_service.get_group(group_id, user_id)

    def _aggregate_currency_totals(self, personal_rows, group_rows):
        totals_map: dict[CurrencyEnum, dict[str, Decimal]] = {}

        for row in personal_rows:
            currency = self._to_currency(row.currency)
            current = totals_map.setdefault(
                currency,
                {
                    "personal_amount": Decimal("0"),
                    "group_amount": Decimal("0"),
                },
            )
            current["personal_amount"] += self._to_decimal(row.total_amount)

        for row in group_rows:
            currency = self._to_currency(row.currency)
            current = totals_map.setdefault(
                currency,
                {
                    "personal_amount": Decimal("0"),
                    "group_amount": Decimal("0"),
                },
            )
            current["group_amount"] += self._to_decimal(row.total_amount)

        own_vs_group = []
        totals_by_currency = []
        for currency in sorted(totals_map.keys(), key=lambda current: current.value):
            personal_amount = totals_map[currency]["personal_amount"]
            group_amount = totals_map[currency]["group_amount"]
            total_amount = personal_amount + group_amount

            own_vs_group.append(
                {
                    "currency": currency,
                    "personal_amount": personal_amount,
                    "group_amount": group_amount,
                    "total_amount": total_amount,
                }
            )
            totals_by_currency.append({"currency": currency, "total_amount": total_amount})

        total_map_only = {
            currency: item["personal_amount"] + item["group_amount"]
            for currency, item in totals_map.items()
        }

        return totals_by_currency, own_vs_group, total_map_only

    @staticmethod
    def _combine_top_categories(personal_rows, group_rows, limit: int):
        category_map: dict[int, dict[str, Decimal | str | int]] = {}

        for row in list(personal_rows) + list(group_rows):
            category_id = int(row.category_id)
            current = category_map.setdefault(
                category_id,
                {
                    "category_id": category_id,
                    "category_name": row.category_name,
                    "total_amount": Decimal("0"),
                },
            )
            current["total_amount"] = current["total_amount"] + Decimal(str(row.total_amount or 0))

        ordered = sorted(
            category_map.values(),
            key=lambda item: (item["total_amount"], str(item["category_name"]).lower()),
            reverse=True,
        )

        return ordered[:limit]

    @staticmethod
    def _combine_top_groups(group_rows, limit: int):
        group_map: dict[int, dict[str, Decimal | str | int]] = {}

        for row in group_rows:
            group_id = int(row.group_id)
            current = group_map.setdefault(
                group_id,
                {
                    "group_id": group_id,
                    "group_name": row.group_name,
                    "total_amount": Decimal("0"),
                },
            )
            current["total_amount"] = current["total_amount"] + Decimal(str(row.total_amount or 0))

        ordered = sorted(
            group_map.values(),
            key=lambda item: (item["total_amount"], str(item["group_name"]).lower()),
            reverse=True,
        )

        return ordered[:limit]

    @staticmethod
    def _build_comparison(current_totals: dict[CurrencyEnum, Decimal], previous_totals: dict[CurrencyEnum, Decimal]):
        comparison = []
        currencies = sorted(
            set(current_totals.keys()) | set(previous_totals.keys()),
            key=lambda current: current.value,
        )

        for currency in currencies:
            current_total = current_totals.get(currency, Decimal("0"))
            previous_total = previous_totals.get(currency, Decimal("0"))
            delta_amount = current_total - previous_total

            if previous_total == 0:
                delta_percent = None
            else:
                delta_percent = float((delta_amount / previous_total) * Decimal("100"))

            comparison.append(
                {
                    "currency": currency,
                    "current_total": current_total,
                    "previous_total": previous_total,
                    "delta_amount": delta_amount,
                    "delta_percent": delta_percent,
                }
            )

        return comparison

    def _build_period_overview(
        self,
        user_id: int,
        period: tuple[date, date],
        scope: SummaryScope,
        category_ids: list[int] | None,
        currency: CurrencyEnum | None,
        group_id: int | None,
        top_categories_limit: int,
        top_groups_limit: int,
    ):
        date_from_dt, date_to_dt = self._to_datetime_range(period)

        personal_count = 0
        personal_totals = []
        personal_categories = []

        group_count = 0
        group_totals = []
        group_categories = []
        group_rows = []

        if scope in ("all", "personal"):
            personal_count = self.expense_repo.get_personal_total_count(
                user_id=user_id,
                date_from=date_from_dt,
                date_to=date_to_dt,
                category_ids=category_ids,
                currency=currency,
            )
            personal_totals = self.expense_repo.get_personal_totals_by_currency(
                user_id=user_id,
                date_from=date_from_dt,
                date_to=date_to_dt,
                category_ids=category_ids,
                currency=currency,
            )
            personal_categories = self.expense_repo.get_personal_top_categories(
                user_id=user_id,
                date_from=date_from_dt,
                date_to=date_to_dt,
                category_ids=category_ids,
                currency=currency,
            )

        if scope in ("all", "group"):
            group_count = self.expense_repo.get_group_share_total_count(
                user_id=user_id,
                date_from=date_from_dt,
                date_to=date_to_dt,
                category_ids=category_ids,
                currency=currency,
                group_id=group_id,
            )
            group_totals = self.expense_repo.get_group_share_totals_by_currency(
                user_id=user_id,
                date_from=date_from_dt,
                date_to=date_to_dt,
                category_ids=category_ids,
                currency=currency,
                group_id=group_id,
            )
            group_categories = self.expense_repo.get_group_share_top_categories(
                user_id=user_id,
                date_from=date_from_dt,
                date_to=date_to_dt,
                category_ids=category_ids,
                currency=currency,
                group_id=group_id,
            )
            group_rows = self.expense_repo.get_group_share_top_groups(
                user_id=user_id,
                date_from=date_from_dt,
                date_to=date_to_dt,
                category_ids=category_ids,
                currency=currency,
                group_id=group_id,
            )

        totals_by_currency, own_vs_group, totals_map = self._aggregate_currency_totals(personal_totals, group_totals)

        return {
            "total_count": personal_count + group_count,
            "totals_by_currency": totals_by_currency,
            "own_vs_group": own_vs_group,
            "top_categories": self._combine_top_categories(personal_categories, group_categories, top_categories_limit),
            "top_groups": self._combine_top_groups(group_rows, top_groups_limit),
            "totals_map": totals_map,
        }

    def get_overview(
        self,
        user_id: int,
        date_from: date | None = None,
        date_to: date | None = None,
        scope: SummaryScope = "all",
        category_id: int | None = None,
        category_ids: list[int] | None = None,
        currency: CurrencyEnum | None = None,
        group_id: int | None = None,
        top_categories_limit: int = 5,
        top_groups_limit: int = 5,
        compare_previous: bool = True,
    ):
        resolved_scope = self._resolve_scope(scope)
        self._validate_scope_group_filter(resolved_scope, group_id, user_id)
        resolved_category_ids = self._resolve_category_ids(category_id=category_id, category_ids=category_ids)

        current_period, previous_period = self._resolve_periods(
            date_from=date_from,
            date_to=date_to,
            compare_previous=compare_previous,
        )

        current = self._build_period_overview(
            user_id=user_id,
            period=current_period,
            scope=resolved_scope,
            category_ids=resolved_category_ids,
            currency=currency,
            group_id=group_id,
            top_categories_limit=top_categories_limit,
            top_groups_limit=top_groups_limit,
        )

        comparison = []
        if previous_period is not None:
            previous = self._build_period_overview(
                user_id=user_id,
                period=previous_period,
                scope=resolved_scope,
                category_ids=resolved_category_ids,
                currency=currency,
                group_id=group_id,
                top_categories_limit=top_categories_limit,
                top_groups_limit=top_groups_limit,
            )
            comparison = self._build_comparison(current["totals_map"], previous["totals_map"])

        return {
            "total_count": current["total_count"],
            "totals_by_currency": current["totals_by_currency"],
            "own_vs_group": current["own_vs_group"],
            "top_categories": current["top_categories"],
            "top_groups": current["top_groups"],
            "comparison_by_currency": comparison,
            "current_period": {
                "date_from": current_period[0],
                "date_to": current_period[1],
            },
            "previous_period": (
                {
                    "date_from": previous_period[0],
                    "date_to": previous_period[1],
                }
                if previous_period is not None
                else None
            ),
        }

    def _collect_period_trends(
        self,
        user_id: int,
        period: tuple[date, date],
        scope: SummaryScope,
        category_ids: list[int] | None,
        currency: CurrencyEnum | None,
        group_id: int | None,
    ):
        date_from_dt, date_to_dt = self._to_datetime_range(period)

        personal_rows = []
        group_rows = []

        if scope in ("all", "personal"):
            personal_rows = self.expense_repo.get_personal_daily_trends(
                user_id=user_id,
                date_from=date_from_dt,
                date_to=date_to_dt,
                category_ids=category_ids,
                currency=currency,
            )

        if scope in ("all", "group"):
            group_rows = self.expense_repo.get_group_share_daily_trends(
                user_id=user_id,
                date_from=date_from_dt,
                date_to=date_to_dt,
                category_ids=category_ids,
                currency=currency,
                group_id=group_id,
            )

        trend_map: dict[CurrencyEnum, dict[date, dict[str, Decimal]]] = {}

        for row in personal_rows:
            currency_key = self._to_currency(row.currency)
            day_key = self._normalize_day(row.day)
            currency_bucket = trend_map.setdefault(currency_key, {})
            day_bucket = currency_bucket.setdefault(
                day_key,
                {"personal_amount": Decimal("0"), "group_amount": Decimal("0")},
            )
            day_bucket["personal_amount"] += self._to_decimal(row.total_amount)

        for row in group_rows:
            currency_key = self._to_currency(row.currency)
            day_key = self._normalize_day(row.day)
            currency_bucket = trend_map.setdefault(currency_key, {})
            day_bucket = currency_bucket.setdefault(
                day_key,
                {"personal_amount": Decimal("0"), "group_amount": Decimal("0")},
            )
            day_bucket["group_amount"] += self._to_decimal(row.total_amount)

        return trend_map

    def get_trends(
        self,
        user_id: int,
        date_from: date | None = None,
        date_to: date | None = None,
        scope: SummaryScope = "all",
        granularity: TrendGranularity = "daily",
        category_id: int | None = None,
        category_ids: list[int] | None = None,
        currency: CurrencyEnum | None = None,
        group_id: int | None = None,
        compare_previous: bool = True,
    ):
        resolved_scope = self._resolve_scope(scope)
        resolved_granularity = self._resolve_trend_granularity(granularity)
        self._validate_scope_group_filter(resolved_scope, group_id, user_id)
        resolved_category_ids = self._resolve_category_ids(category_id=category_id, category_ids=category_ids)

        current_period, previous_period = self._resolve_periods(
            date_from=date_from,
            date_to=date_to,
            compare_previous=compare_previous,
        )

        current_map = self._collect_period_trends(
            user_id=user_id,
            period=current_period,
            scope=resolved_scope,
            category_ids=resolved_category_ids,
            currency=currency,
            group_id=group_id,
        )

        previous_map: dict[CurrencyEnum, dict[date, dict[str, Decimal]]] = {}
        if previous_period is not None:
            previous_map = self._collect_period_trends(
                user_id=user_id,
                period=previous_period,
                scope=resolved_scope,
                category_ids=resolved_category_ids,
                currency=currency,
                group_id=group_id,
            )

        current_map = self._aggregate_trend_buckets(
            trend_map=current_map,
            period=current_period,
            granularity=resolved_granularity,
        )
        previous_map = self._aggregate_trend_buckets(
            trend_map=previous_map,
            period=previous_period if previous_period is not None else current_period,
            granularity=resolved_granularity,
        )

        currencies = sorted(
            set(current_map.keys()) | set(previous_map.keys()),
            key=lambda current: current.value,
        )
        if not currencies and currency is not None:
            currencies = [currency]

        current_days = list(self._iter_trend_buckets(current_period[0], current_period[1], resolved_granularity))
        previous_days = (
            list(self._iter_trend_buckets(previous_period[0], previous_period[1], resolved_granularity))
            if previous_period
            else []
        )

        currency_series = []
        for currency_key in currencies:
            current_points = []
            for current_day in current_days:
                amounts = current_map.get(currency_key, {}).get(
                    current_day,
                    {"personal_amount": Decimal("0"), "group_amount": Decimal("0")},
                )
                personal_amount = amounts["personal_amount"]
                group_amount = amounts["group_amount"]
                current_points.append(
                    {
                        "date": current_day,
                        "personal_amount": personal_amount,
                        "group_amount": group_amount,
                        "total_amount": personal_amount + group_amount,
                    }
                )

            previous_points = []
            for previous_day in previous_days:
                amounts = previous_map.get(currency_key, {}).get(
                    previous_day,
                    {"personal_amount": Decimal("0"), "group_amount": Decimal("0")},
                )
                personal_amount = amounts["personal_amount"]
                group_amount = amounts["group_amount"]
                previous_points.append(
                    {
                        "date": previous_day,
                        "personal_amount": personal_amount,
                        "group_amount": group_amount,
                        "total_amount": personal_amount + group_amount,
                    }
                )

            currency_series.append(
                {
                    "currency": currency_key,
                    "current": current_points,
                    "previous": previous_points,
                }
            )

        return {
            "current_period": {
                "date_from": current_period[0],
                "date_to": current_period[1],
            },
            "previous_period": (
                {
                    "date_from": previous_period[0],
                    "date_to": previous_period[1],
                }
                if previous_period is not None
                else None
            ),
            "currencies": currency_series,
        }

    def _collect_drilldown_items(
        self,
        user_id: int,
        date_from: date | None,
        date_to: date | None,
        scope: SummaryScope,
        category_ids: list[int] | None,
        currency: CurrencyEnum | None,
        group_id: int | None,
        sort_by: DrilldownSortBy,
        sort_order: SortOrder,
    ):
        date_from_dt = datetime.combine(date_from, time.min) if date_from else None
        date_to_dt = datetime.combine(date_to, time.max) if date_to else None

        rows = []
        if scope in ("all", "personal"):
            rows.extend(
                self.expense_repo.get_personal_summary_records(
                    user_id=user_id,
                    date_from=date_from_dt,
                    date_to=date_to_dt,
                    category_ids=category_ids,
                    currency=currency,
                )
            )

        if scope in ("all", "group"):
            rows.extend(
                self.expense_repo.get_group_share_summary_records(
                    user_id=user_id,
                    date_from=date_from_dt,
                    date_to=date_to_dt,
                    category_ids=category_ids,
                    currency=currency,
                    group_id=group_id,
                )
            )

        items = []
        for row in rows:
            items.append(
                {
                    "expense_id": row.expense_id,
                    "scope": row.scope,
                    "title": row.title,
                    "expense_date": row.expense_date,
                    "created_at": row.created_at,
                    "currency": self._to_currency(row.currency),
                    "category_id": row.category_id,
                    "category_name": row.category_name,
                    "group_id": row.group_id,
                    "group_name": row.group_name,
                    "total_amount": self._to_decimal(row.total_amount),
                    "user_amount": self._to_decimal(row.user_amount),
                    "recurring_expense_id": row.recurring_expense_id,
                    "recurring_occurrence_date": row.recurring_occurrence_date,
                }
            )

        reverse = sort_order == "desc"
        if sort_by == "amount":
            items.sort(key=lambda item: (item["user_amount"], item["expense_id"]), reverse=reverse)
        elif sort_by == "created_at":
            items.sort(key=lambda item: (item["created_at"], item["expense_id"]), reverse=reverse)
        else:
            items.sort(key=lambda item: (item["expense_date"], item["expense_id"]), reverse=reverse)

        return items

    def get_drilldown(
        self,
        user_id: int,
        limit: int,
        offset: int,
        date_from: date | None = None,
        date_to: date | None = None,
        scope: SummaryScope = "all",
        category_id: int | None = None,
        category_ids: list[int] | None = None,
        currency: CurrencyEnum | None = None,
        group_id: int | None = None,
        sort_by: DrilldownSortBy = "expense_date",
        sort_order: SortOrder = "desc",
    ):
        if date_from and date_to and date_from > date_to:
            raise HTTPException(status_code=400, detail="date_from cannot be greater than date_to")

        resolved_scope = self._resolve_scope(scope)
        self._validate_scope_group_filter(resolved_scope, group_id, user_id)
        resolved_category_ids = self._resolve_category_ids(category_id=category_id, category_ids=category_ids)

        items = self._collect_drilldown_items(
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
            scope=resolved_scope,
            category_ids=resolved_category_ids,
            currency=currency,
            group_id=group_id,
            sort_by=sort_by,
            sort_order=sort_order,
        )

        total_count = len(items)
        paginated = items[offset : offset + limit]

        return {
            "total_count": total_count,
            "items": paginated,
        }

    @staticmethod
    def _format_enum_value(value) -> str:
        if hasattr(value, "value"):
            return str(value.value)
        return str(value)

    def _parse_requested_export_sections(self, sections: str | None) -> list[ExportSection]:
        if not sections:
            return []

        raw_items = [item.strip() for item in sections.split(",") if item.strip()]
        if not raw_items:
            return []

        allowed = set(EXPORT_ALL_SECTIONS)
        invalid = [item for item in raw_items if item not in allowed]
        if invalid:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid export section(s): {', '.join(invalid)}",
            )

        unique_items: list[ExportSection] = []
        for item in raw_items:
            if item not in unique_items:
                unique_items.append(item)  # type: ignore[arg-type]

        return unique_items

    def _resolve_export_sections(self, export_format: ExportFormat, sections: str | None) -> list[ExportSection]:
        requested = self._parse_requested_export_sections(sections)
        allowed = EXPORT_ALLOWED_SECTIONS[export_format]

        filtered = [section for section in requested if section in allowed]
        if filtered:
            return filtered

        return list(EXPORT_DEFAULT_SECTIONS[export_format])

    def _collect_category_summary_rows(
        self,
        user_id: int,
        date_from: date | None,
        date_to: date | None,
        scope: SummaryScope,
        category_id: int | None,
        category_ids: list[int] | None,
        currency: CurrencyEnum | None,
        group_id: int | None,
    ) -> list[list[str]]:
        overview = self.get_overview(
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
            scope=scope,
            category_id=category_id,
            category_ids=category_ids,
            currency=currency,
            group_id=group_id,
            top_categories_limit=10,
            top_groups_limit=5,
            compare_previous=False,
        )

        rows: list[list[str]] = []
        for total in overview["totals_by_currency"]:
            rows.append(
                [
                    "total_by_currency",
                    "all",
                    self._format_export_currency(total["currency"]),
                    f"{self._to_decimal(total['total_amount']):.2f}",
                ]
            )

        for category_row in overview["top_categories"]:
            rows.append(
                [
                    "top_category",
                    str(category_row["category_name"]),
                    self._format_export_currency(currency) if currency is not None else "multi",
                    f"{self._to_decimal(category_row['total_amount']):.2f}",
                ]
            )

        return rows

    def export_csv(
        self,
        user_id: int,
        date_from: date | None = None,
        date_to: date | None = None,
        scope: SummaryScope = "all",
        category_id: int | None = None,
        category_ids: list[int] | None = None,
        currency: CurrencyEnum | None = None,
        group_id: int | None = None,
        sections: str | None = None,
        locale: str | None = None,
        filename: str | None = None,
        sort_by: DrilldownSortBy = "expense_date",
        sort_order: SortOrder = "desc",
    ):
        locale_code, localization = self._get_export_localization(locale)
        resolved_sections = self._resolve_export_sections("csv", sections)
        output = io.StringIO(newline="")
        writer = csv.writer(output)

        for index, section in enumerate(resolved_sections):
            if index > 0:
                writer.writerow([])

            if section == "transactions":
                writer.writerow([str(localization["sheet_title"])])
                writer.writerow(CSV_EXPORT_COLUMNS)
                rows = self._prepare_export_rows(
                    user_id=user_id,
                    date_from=date_from,
                    date_to=date_to,
                    scope=scope,
                    category_id=category_id,
                    category_ids=category_ids,
                    currency=currency,
                    group_id=group_id,
                    sort_by=sort_by,
                    sort_order=sort_order,
                )
                writer.writerows(rows)
                continue

            if section == "category_summary":
                writer.writerow([str(localization["category_breakdown"])])
                writer.writerow([str(localization["metric"]), "name", "currency", str(localization["amount"])])
                rows = self._collect_category_summary_rows(
                    user_id=user_id,
                    date_from=date_from,
                    date_to=date_to,
                    scope=scope,
                    category_id=category_id,
                    category_ids=category_ids,
                    currency=currency,
                    group_id=group_id,
                )
                writer.writerows(rows)
                continue

        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        if filename:
            if not filename.endswith(".csv"):
                filename = f"{filename}.csv"
        else:
            filename = f"{localization['filename_prefix']}-{timestamp}.csv"

        return {
            "filename": filename,
            "content": output.getvalue(),
        }

    def export_xlsx(
        self,
        user_id: int,
        date_from: date | None = None,
        date_to: date | None = None,
        scope: SummaryScope = "all",
        category_id: int | None = None,
        category_ids: list[int] | None = None,
        currency: CurrencyEnum | None = None,
        group_id: int | None = None,
        sections: str | None = None,
        locale: str | None = None,
        filename: str | None = None,
        sort_by: DrilldownSortBy = "expense_date",
        sort_order: SortOrder = "desc",
    ):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as error:
            raise HTTPException(status_code=500, detail="XLSX export dependency is not installed") from error

        locale_code, localization = self._get_export_localization(locale)
        resolved_sections = self._resolve_export_sections("xlsx", sections)

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        header_fill = PatternFill(fill_type="solid", start_color="0F172A", end_color="0F172A")
        header_font = Font(color="FFFFFF", bold=True)

        def add_sheet(title: str, columns: list[str], rows: list[list[str]], column_widths: list[int] | None = None):
            sheet_title = title[:31] if title else "Sheet"
            worksheet = workbook.create_sheet(title=sheet_title)
            worksheet.append(columns)
            for row in rows:
                worksheet.append(row)

            worksheet.freeze_panes = "A2"
            for col_index, _ in enumerate(columns, start=1):
                cell = worksheet.cell(row=1, column=col_index)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if column_widths is None:
                column_widths = [max(12, len(column) + 2) for column in columns]

            for col_index, width in enumerate(column_widths, start=1):
                worksheet.column_dimensions[get_column_letter(col_index)].width = width

        if "transactions" in resolved_sections:
            _, _, transaction_rows = self._prepare_user_friendly_export_rows(
                user_id=user_id,
                date_from=date_from,
                date_to=date_to,
                scope=scope,
                category_id=category_id,
                category_ids=category_ids,
                currency=currency,
                group_id=group_id,
                locale=locale,
                sort_by=sort_by,
                sort_order=sort_order,
            )
            add_sheet(
                title=str(localization["sheet_title"]),
                columns=[str(column) for column in localization["columns"]],
                rows=transaction_rows,
                column_widths=[14, 34, 14, 24, 24, 12, 16, 16],
            )

        if "category_summary" in resolved_sections:
            category_rows = self._collect_category_summary_rows(
                user_id=user_id,
                date_from=date_from,
                date_to=date_to,
                scope=scope,
                category_id=category_id,
                category_ids=category_ids,
                currency=currency,
                group_id=group_id,
            )
            add_sheet(
                title=str(localization["category_breakdown"]),
                columns=[str(localization["metric"]), "Name", "Currency", str(localization["amount"])],
                rows=category_rows,
            )

        if len(workbook.sheetnames) == 0:
            add_sheet(title="Summary", columns=["Info"], rows=[["No sections selected"]])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        if filename:
            if not filename.endswith(".xlsx"):
                filename = f"{filename}.xlsx"
        else:
            filename = f"{localization['filename_prefix']}-{timestamp}.xlsx"

        return {
            "filename": filename,
            "content": output.getvalue(),
        }

    def export_pdf(
        self,
        user_id: int,
        date_from: date | None = None,
        date_to: date | None = None,
        scope: SummaryScope = "all",
        category_id: int | None = None,
        category_ids: list[int] | None = None,
        currency: CurrencyEnum | None = None,
        group_id: int | None = None,
        sections: str | None = None,
        locale: str | None = None,
        filename: str | None = None,
        sort_by: DrilldownSortBy = "expense_date",
        sort_order: SortOrder = "desc",
    ):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.pdfbase.cidfonts import CIDFont
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            # Register font with Polish character support
            # Try to use system fonts that support Polish characters
            import os
            import platform

            font_name = 'Helvetica'
            font_bold_name = 'Helvetica-Bold'

            try:
                if platform.system() == 'Windows':
                    # Try to use Arial on Windows (better Unicode support)
                    arial_path = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts', 'arial.ttf')
                    arial_bold_path = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts', 'arialbd.ttf')
                    if os.path.exists(arial_path) and os.path.exists(arial_bold_path):
                        pdfmetrics.registerFont(TTFont('Arial', arial_path, subfontIndex=0))
                        pdfmetrics.registerFont(TTFont('Arial-Bold', arial_bold_path, subfontIndex=0))
                        font_name = 'Arial'
                        font_bold_name = 'Arial-Bold'
            except Exception as e:
                # Fallback to Helvetica if system fonts are not available
                pass
        except ImportError as error:
            raise HTTPException(status_code=500, detail="PDF export dependency is not installed") from error

        resolved_sections = self._resolve_export_sections("pdf", sections)
        locale_code, localization = self._get_export_localization(locale)
        overview = self.get_overview(
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
            scope=scope,
            category_id=category_id,
            category_ids=category_ids,
            currency=currency,
            group_id=group_id,
            top_categories_limit=8,
            top_groups_limit=5,
            compare_previous=False,
        )

        pdf_output = io.BytesIO()
        document = SimpleDocTemplate(
            pdf_output,
            pagesize=landscape(A4),
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
            encoding='utf-8',
        )

        styles = getSampleStyleSheet()
        # Update specific styles to use registered font
        styles["Heading3"].fontName = font_name
        styles["Heading4"].fontName = font_name
        styles["Normal"].fontName = font_name

        generated_at = self._format_export_datetime(datetime.now(), locale_code)
        document_title = str(localization["document_title"])
        generated_label = str(localization["generated_at"])

        story = [
            Paragraph(document_title, styles["Heading3"]),
            Paragraph(f"{generated_label}: {generated_at}", styles["Normal"]),
            Spacer(1, 8),
        ]

        kpi_rows = [[str(localization["metric"]), str(localization["value"])], [str(localization["total_transactions"]), str(overview["total_count"])]]
        for total_row in overview["totals_by_currency"]:
            kpi_rows.append(
                [
                    f"Total ({self._format_export_currency(total_row['currency'])})",
                    f"{self._to_decimal(total_row['total_amount']):.2f}",
                ]
            )

        kpi_table = Table(kpi_rows, repeatRows=1, colWidths=[90 * mm, 70 * mm])
        kpi_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), font_bold_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ]
            )
        )
        story.extend([Paragraph(str(localization["kpi_summary"]), styles["Heading4"]), kpi_table, Spacer(1, 10)])

        if "category_summary" in resolved_sections:
            category_rows = [[str(localization["category"]), str(localization["amount"])]]
            for category_row in overview["top_categories"]:
                category_name = category_row["category_name"]
                # Try to use localized category name, fallback to original name
                localized_name = localization["categories"].get(category_name, category_name)
                category_rows.append(
                    [
                        localized_name,
                        f"{self._to_decimal(category_row['total_amount']):.2f}",
                    ]
                )

            category_table = Table(category_rows, repeatRows=1, colWidths=[120 * mm, 40 * mm])
            category_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), font_bold_name),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ]
                )
            )
            story.extend([Paragraph(str(localization["category_breakdown"]), styles["Heading4"]), category_table, Spacer(1, 10)])

        if "transactions" in resolved_sections:
            transaction_rows = self._prepare_export_rows(
                user_id=user_id,
                date_from=date_from,
                date_to=date_to,
                scope=scope,
                category_id=category_id,
                category_ids=category_ids,
                currency=currency,
                group_id=group_id,
                sort_by=sort_by,
                sort_order=sort_order,
            )

            if transaction_rows:
                header_row = [str(col) for col in localization["columns"]]
                table_data = [header_row]
                for row in transaction_rows:
                    # Map CSV columns to display columns: expense_date(3), title(2), scope(1), category_name(7), group_name(9), currency(5), total_amount(10), user_amount(11)
                    category_name = row[7]
                    # Try to use localized category name, fallback to original name
                    localized_category = localization["categories"].get(category_name, category_name)
                    display_row = [
                        self._format_export_date(datetime.fromisoformat(row[3]), locale_code),
                        row[2],  # title
                        localization["scope"].get(row[1], row[1]),  # scope
                        localized_category,  # category_name (localized)
                        row[9] if row[9] else localization["empty_group"],  # group_name
                        row[5],  # currency
                        row[10],  # total_amount
                        row[11],  # user_amount
                    ]
                    table_data.append(display_row)

                transaction_table = Table(table_data, repeatRows=1, colWidths=[20 * mm, 50 * mm, 20 * mm, 30 * mm, 30 * mm, 15 * mm, 25 * mm, 25 * mm])
                transaction_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("FONTNAME", (0, 0), (-1, 0), font_bold_name),
                            ("FONTSIZE", (0, 0), (-1, -1), 7),
                            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                        ]
                    )
                )
                story.extend([Spacer(1, 5), Paragraph(str(localization["sheet_title"]), styles["Heading4"]), transaction_table, Spacer(1, 10)])

        document.build(
            story
        )

        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        if filename:
            if not filename.endswith(".pdf"):
                filename = f"{filename}.pdf"
        else:
            filename = f"{localization['filename_prefix']}-{timestamp}.pdf"

        return {
            "filename": filename,
            "content": pdf_output.getvalue(),
        }

    def _get_export_items(
        self,
        user_id: int,
        date_from: date | None,
        date_to: date | None,
        scope: SummaryScope,
        category_id: int | None,
        category_ids: list[int] | None,
        currency: CurrencyEnum | None,
        group_id: int | None,
        sort_by: DrilldownSortBy,
        sort_order: SortOrder,
    ):
        if date_from and date_to and date_from > date_to:
            raise HTTPException(status_code=400, detail="date_from cannot be greater than date_to")

        resolved_scope = self._resolve_scope(scope)
        self._validate_scope_group_filter(resolved_scope, group_id, user_id)
        resolved_category_ids = self._resolve_category_ids(category_id=category_id, category_ids=category_ids)

        return self._collect_drilldown_items(
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
            scope=resolved_scope,
            category_ids=resolved_category_ids,
            currency=currency,
            group_id=group_id,
            sort_by=sort_by,
            sort_order=sort_order,
        )

    def _prepare_user_friendly_export_rows(
        self,
        user_id: int,
        date_from: date | None,
        date_to: date | None,
        scope: SummaryScope,
        category_id: int | None,
        category_ids: list[int] | None,
        currency: CurrencyEnum | None,
        group_id: int | None,
        locale: str | None,
        sort_by: DrilldownSortBy,
        sort_order: SortOrder,
    ):
        locale_code, localization = self._get_export_localization(locale)
        scope_labels = localization["scope"]
        empty_group = str(localization["empty_group"])

        items = self._get_export_items(
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
            scope=scope,
            category_id=category_id,
            category_ids=category_ids,
            currency=currency,
            group_id=group_id,
            sort_by=sort_by,
            sort_order=sort_order,
        )

        rows: list[list[str]] = []
        for item in items:
            scope_label = str(scope_labels["personal"]) if item["scope"] == "personal" else str(scope_labels["group"])
            rows.append(
                [
                    self._format_export_date(item["expense_date"], locale_code),
                    str(item["title"]),
                    scope_label,
                    str(item["category_name"]),
                    str(item["group_name"]) if item["group_name"] is not None else empty_group,
                    self._format_export_currency(item["currency"]),
                    self._format_export_amount(self._to_decimal(item["total_amount"]), locale_code),
                    self._format_export_amount(self._to_decimal(item["user_amount"]), locale_code),
                ]
            )

        return locale_code, localization, rows

    def _prepare_export_rows(
        self,
        user_id: int,
        date_from: date | None,
        date_to: date | None,
        scope: SummaryScope,
        category_id: int | None,
        category_ids: list[int] | None,
        currency: CurrencyEnum | None,
        group_id: int | None,
        sort_by: DrilldownSortBy,
        sort_order: SortOrder,
    ) -> list[list[str]]:
        items = self._get_export_items(
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
            scope=scope,
            category_id=category_id,
            category_ids=category_ids,
            currency=currency,
            group_id=group_id,
            sort_by=sort_by,
            sort_order=sort_order,
        )

        rows: list[list[str]] = []
        for item in items:
            rows.append(
                [
                    str(item["expense_id"]),
                    str(item["scope"]),
                    str(item["title"]),
                    item["expense_date"].isoformat(),
                    item["created_at"].isoformat(),
                    self._format_export_currency(item["currency"]),
                    str(item["category_id"]),
                    str(item["category_name"]),
                    str(item["group_id"]) if item["group_id"] is not None else "",
                    str(item["group_name"]) if item["group_name"] is not None else "",
                    f"{Decimal(item['total_amount']):.2f}",
                    f"{Decimal(item['user_amount']):.2f}",
                    str(item["recurring_expense_id"]) if item["recurring_expense_id"] is not None else "",
                    item["recurring_occurrence_date"].isoformat() if item["recurring_occurrence_date"] is not None else "",
                ]
            )

        return rows
